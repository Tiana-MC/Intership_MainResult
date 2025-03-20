"""
Created on 10:30 6/13/2024
Author  : <TianMC>
"""
import pandas as pd
import numpy as np
from WindPy import w
import warnings
import get_data as gdt
import statsmodels.api as sm
import cvxpy as cp
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
w.start()
w.isconnected()
warnings.filterwarnings('ignore')

# ----------------------------初始化----------------
# 变量设置
start_date = '2014-01-01'  # 若已提取数据，则该变量几乎没啥用，因为数据已提取，只会根据end_date更新数据
end_date = '2024-06-13'

used_data_path = r'\data\used_data'
result_file_path = r'\data\ret_result'
fund_weight_data_path = r'\data\fund_stock_weight'
bench_weight_data_path = r'\data\bench_stock_weight'
daily_ret_data_path = r'\data\fund_bench_daily_ret'
semiyear_ret_data_path = r'\data\fund_bench_semiyear_ret'

current_path = os.path.dirname(os.path.abspath(__file__))  # 获取当前脚本所在的路径
used_data_path = current_path + used_data_path
result_file_path = current_path + result_file_path
fund_weight_data_path = current_path + fund_weight_data_path
bench_weight_data_path = current_path + bench_weight_data_path
daily_ret_data_path = current_path + daily_ret_data_path
semiyear_ret_data_path = current_path + semiyear_ret_data_path


# region 通用数据读取
# 股票行业数据
stocks_ind = gdt.get_stock_ind(start_dt=start_date, end_dt=end_date, struc_path=used_data_path,
                               ind_path=used_data_path)

# 基准每日收益率
bench_list = ['000300.SH', '000905.SH', '000852.SH', '000922.CSI', '399006.SZ']
bench_ret = {}
for bench in bench_list:
    bench_ret[bench] = gdt.get_ret_daily(start_dt=start_date, end_dt=end_date, wcode=bench,
                                         file_path=daily_ret_data_path, bench=True) / 100

# 股票收益率数据
stocks_ret = gdt.get_stock_ret(start_dt=start_date, end_dt=end_date, struc_path=used_data_path,
                               ret_path=used_data_path) / 100
# endregion

all_ind = ['银行', '房地产', '医药', '电力及公用事业', '机械', '综合', '建筑', '建材', '家电', '食品饮料', '电子', '汽车',
           '商贸零售', '通信', '传媒', '农林牧渔', '石油石化', '有色金属', '计算机', '交通运输', '基础化工', '非银行金融',
           '电力设备及新能源', '消费者服务', '轻工制造', '国防军工', '煤炭', '纺织服装', '钢铁', '综合金融']  # 30个中信一级行业
# -------------------------------------------------------------------


def solver(X, y):
    """
    :param X:  矩阵,shape(xx,n)
    :param y:  矩阵,shape(xx,)，和 X的shape[0]相等
    :return: 列表，X的系数
    """
    # 针对：太新的日期，基金的每日收益数据为nan
    min_len = min(len(X), len(y))
    X = X[:min_len]
    y = y[:min_len]

    # 未知数的个数
    n = X.shape[1]

    # 定义优化变量
    a = cp.Variable(n)
    b = cp.Variable()

    # 定义目标函数
    residuals = y - (X @ a + b)
    objective = cp.Minimize(cp.sum_squares(residuals))

    # 定义约束条件
    constraints = [
        a >= 0,
        cp.sum(a) <= 0.95
    ]

    # 定义优化问题
    problem = cp.Problem(objective, constraints)

    # 求解问题
    problem.solve()
    return a.value


def multi_beta(study_fund, fund_ret, bench_ret):
    n = len(study_fund.index.unique())

    # 沪深300，中证500，中证1000回归, 000300.SH, 000905.SH, 000852.SH
    beta_df1 = pd.DataFrame(index=np.arange(1, n+1), columns=['start_dt', 'end_dt', '000300.SH', '000905.SH', '000852.SH'])

    # 中证红利，创业板指，000922.CSI， 399006.SZ
    beta_df2 = pd.DataFrame(index=np.arange(1, n+1), columns=['start_dt', 'end_dt', '000922.CSI', '399006.SZ'])

    for i in range(1, n+1):  # 期数，1到n+1
        X1 = pd.DataFrame(columns=beta_df1.columns[2:])
        X2 = pd.DataFrame(columns=beta_df2.columns[2:])

        start_i, end_i = period_date(study_fund).loc[i]

        # 遍历字典并提取指定区间的数据
        for key, bench in bench_ret.items():
            if key in X1.columns:
                X1[key] = bench.loc[start_i:end_i]
            if key in X2.columns:
                X2[key] = bench.loc[start_i:end_i]

        X1 = np.array(X1)
        X2 = np.array(X2)
        y = fund_ret.loc[start_i:end_i].values.reshape(-1)

        beta_df1.iloc[i-1, :2] = [start_i, end_i]
        beta_df1.iloc[i-1, 2:] = solver(X1, y)

        beta_df2.iloc[i-1, :2] = [start_i, end_i]
        beta_df2.iloc[i-1, 2:] = solver(X2, y)
    return beta_df1, beta_df2


# 基金风格估计，基金收益率对000300.SH 000922.CSI 000905.SH 000852.SH 399006.SZ 收益率回归的R2
def regression_r2(study_fund, fund_ret, bench_ret):
    n = len(study_fund.index.unique())
    r2_df = pd.DataFrame(index=np.arange(1, n+1), columns=['start_dt', 'end_dt', '000300.SH', '000905.SH', '000852.SH', '000922.CSI', '399006.SZ'])
    for i in range(1, n+1):  # 期数，1到n+1
        start_i, end_i = period_date(study_fund).loc[i]
        r2_i = pd.DataFrame(index=['r2'], columns=['000300.SH', '000905.SH', '000852.SH', '000922.CSI', '399006.SZ'])

        # 对每个基准都进行回归
        bench_list = bench_ret.keys()
        for bench in bench_list:
            X = bench_ret[bench].loc[start_i:end_i].values
            X = sm.add_constant(X)
            y = fund_ret.loc[start_i:end_i].values

            # 针对：太新的日期，基金的每日收益数据为nan
            if len(X) != len(y):
                X = X[: len(y)]

            model = sm.OLS(y, X).fit()
            r2_i[bench] = model.rsquared
        r2_df.iloc[i-1, :2] = [start_i, end_i]
        r2_df.iloc[i-1, 2:] = r2_i.loc['r2'].values
    return r2_df


# 获取start_dt, end_dt的df
def period_date(fund_weight):
    date_list = fund_weight.index.get_level_values(0).unique().tolist()
    date = pd.DataFrame(index=np.arange(1, len(date_list) + 1), columns=['start_date', 'end_date'])
    for i in date.index:
        start_dt = date_list[i - 1]
        try:
            end_dt = date_list[i]
        except:
            global end_date
            end_dt = end_date
        date.loc[i] = [start_dt, end_dt]
    date['start_date'] = pd.to_datetime(date['start_date'])
    date['end_date'] = pd.to_datetime(date['end_date'])
    return date


def singleP(study_fund, bench_fund, stock_ret):
    """
    :param study_fund: stcode, weight, indname三列
    :param bench_fund: stcode, weight, indname三列
    :param stock_ret: 单期区间内的股票收益
    :return: [ER, AR, SR, IR, return_fund, return_bench]
    """
    # study_fund = study_fund_now
    # bench_fund = bench_fund_now
    # stock_ret = stock_ret

    global all_ind

    # 匹配基金持仓股票的所属行业和收益率
    fund = pd.merge(study_fund, stock_ret, on='stcode')  # 匹配每支持仓股票申万一级行业数据,可能有些股票不在这个全部A股及其行业.xlsx文件里，省略不计
    fund['return'] = fund['weight'] * fund['stk_return']

    # 基金各个行业权重,fund_weight:cols=['industry', 'ind_weight']
    fund_group_w = fund.groupby('indname')['weight'].sum().to_frame('ind_weight')
    fund_weight = pd.DataFrame(index=all_ind)
    fund_weight = pd.concat([fund_weight, fund_group_w], axis=1).reset_index().rename(columns={'index': 'industry'})

    # 基金各个行业收益率,fund_return, cols=['industry', 'ind_return']
    fund_group_r = (fund.groupby('indname')['return'].sum()) / (fund.groupby('indname')['weight'].sum())
    fund_return = pd.DataFrame(index=all_ind)
    fund_return = pd.concat([fund_return, fund_group_r], axis=1).reset_index().rename(
        columns={'index': 'industry', 0: 'ind_return'})

    # 基金的各资产权重及收益率
    # print(fund_return, fund_weight)
    fund_weight_return = pd.merge(fund_return, fund_weight, on='industry').fillna(0)


    # 匹配基准的持仓股票的收益率
    bench = pd.merge(bench_fund, stock_ret, on='stcode')
    bench['return'] = bench['weight'] * bench['stk_return']

    # 基准各个行业权重
    bench_group_w = bench.groupby('indname')['weight'].sum().to_frame('ind_weight')
    bench_weight = pd.DataFrame(index=all_ind)
    bench_weight = pd.concat([bench_weight, bench_group_w], axis=1).reset_index().rename(columns={'index': 'industry'})

    # 基准各个行业收益率
    bench_group_r = (bench.groupby('indname')['return'].sum()) / (bench.groupby('indname')['weight'].sum())
    bench_return = pd.DataFrame(index=all_ind)
    bench_return = pd.concat([bench_return, bench_group_r], axis=1).reset_index().rename(
        columns={'index': 'industry', 0: 'ind_return'})

    # 基准的各资产权重及收益率
    # print(bench_return, bench_weight)
    bench_weight_return = pd.merge(bench_return, bench_weight, on='industry').fillna(0)

    # Brinson模型部分
    return_fund = sum(fund_weight_return['ind_weight'] * fund_weight_return['ind_return'])
    return_bench = sum(bench_weight_return['ind_weight'] * bench_weight_return['ind_return'])

    # BF分解法
    ER = return_fund - return_bench
    AR = sum((fund_weight_return['ind_weight'] - bench_weight_return['ind_weight']) * bench_weight_return['ind_return'])
    SR = sum((fund_weight_return['ind_return'] - bench_weight_return['ind_return']) * bench_weight_return['ind_weight'])
    IR = ER - AR - SR
    return [ER, AR, SR, IR, return_fund, return_bench]


# 未经修正的多期，就是每个单期的结果拼成一个dataframe
def mulitP_ori(study_fund, bench_fund, stocks_ret, fund_ret, index_ret):
    """
    :param fund_code: 基金wcode
    :param bench_code: 基准wcode
    :return: dataframe，index为np.arange(1, n+1)，第几期；columns是每期的['ER', 'AR', 'SR', 'IR', 'return_fund_estimate', 'return_bench_estimate', 'return_fund_real', 'return_bench_real', 'fund_return_management']
    """
    n = len(study_fund.index.unique())
    result = pd.DataFrame(index=np.arange(1, n+1), columns=['start_dt', 'end_dt', 'ER', 'AR', 'SR', 'IR',
                                                            'fund_estimate', 'bench_estimate', 'fund_real',
                                                            'bench_real', 'fund_management'])
    for i in range(1, n+1):  # 期数，1到n+1
        start_i, end_i = period_date(study_fund).loc[i].to_list()

        # 基金数据 ['stcode', 'weight', 'indname'] 3列
        study_fund_now = study_fund.loc[start_i].reset_index(drop=True)   # 取中间日期的数据,去掉日期索引，['stcode', 'weight', 'indname'] 3列

        # 基准数据  ['stcode', 'weight', 'indname'] 3列
        bench_fund_now = bench_fund.loc[start_i].reset_index(drop=True)

        # 股票收益率 index为stcode, col为stk_return
        stock_ret_i = stocks_ret.loc[start_i]
        stock_ret_i.name = 'stk_return'
        stock_ret_now = pd.DataFrame(stock_ret_i)
        stock_ret_now.index.name = 'stcode'

        # print(study_fund_now,'\n', bench_fund_now, '\n', stock_ret_now)

        # 单期结果
        ERi, ARi, SRi, IRi, return_fund_i, return_bench_i = iter(singleP(study_fund_now, bench_fund_now, stock_ret_now))
        # print(ERi, ARi, SRi, IRi, return_fund_i, return_bench_i)

        # 基金区间内的实际收益
        real_ret_i = fund_ret.loc[start_i].values[0]
        real_manage_i = real_ret_i - return_fund_i

        # 基准区间内的实际收益
        real_bench_i = index_ret.loc[start_i].values[0]
        result.loc[i] = [start_i, end_i, ERi, ARi, SRi, IRi, return_fund_i, return_bench_i, real_ret_i, real_bench_i, real_manage_i]
    # result['start_dt'] = pd.to_datetime(result['start_dt'])
    # result['end_dt'] = pd.to_datetime(result['end_dt'])
    return result


# 时间区间内的基金换手率
def exchange_rate(study_fund, stocks_ret):
    """
    :return: index为date, col为exchange_rate的换手率
    """
    # shape multi-index fund_weight (index=['date', 'stcode'], col=['weight'])
    fund_weight = study_fund.reset_index().set_index(['date', 'stcode'])[['weight']]

    # add cash and its weight in the fund_weight
    cash_values = 1 - fund_weight.groupby(level='date')['weight'].sum()
    cash_df = pd.DataFrame({'date': cash_values.index, 'stcode': 'cash', 'weight': cash_values.values}).set_index(['date', 'stcode'])  # 创建index和fund_weight相同的，col为cash的weight的df
    fund_weight = pd.concat([fund_weight, cash_df]).sort_values(by=['date', 'stcode'])

    # add the return of the selected data
    tmp = fund_weight.unstack()
    tmp.columns = [col[1] for col in tmp.columns]  # 取消unstack之后的双重col，只保留stcode
    
    # add the return of the selected data
    tmp = fund_weight.unstack()
    tmp.columns = [col[1] for col in tmp.columns]
    fund_weight['return'] = (tmp * stocks_ret).stack().rename_axis(['date', 'stcode'])

    # get the weight_end of stock
    fund_weight['weight_end'] = fund_weight['weight'] * (1 + fund_weight['return'])

    # normalize the end_weight
    fund_weight['weight_end_nor'] = fund_weight['weight_end'] / fund_weight.groupby('date')[
        'weight_end'].transform('sum')

    # exchange_df, exchange_rate of t0  = sum(abs(t0_weight_end_nor - t1_weight))
    exchange_df = pd.DataFrame(index=fund_weight.index.get_level_values(0).unique(), columns=['exchange_rate', 'stock_position'])
    exchange_df.index.name = 'start_dt'
    for idx in exchange_df.index:
        if idx == exchange_df.index[-1]:
            exchange_df.loc[idx] = np.nan
        else:
            # the end_weight_nor of idx
            t0_end_weight_nor = fund_weight.loc[idx]['weight_end_nor']

            # the start_weight of idx+1
            t1_start_weight = fund_weight.loc[exchange_df.index[exchange_df.index.get_loc(idx) + 1]]['weight']

            # exchange_rate of idx
            exchange_df.loc[idx, 'exchange_rate'] = sum(abs(t0_end_weight_nor.sub(t1_start_weight, fill_value=0)))
            exchange_df.loc[idx, 'stock_position'] = fund_weight.loc[idx]['weight_end'].fillna(0).sum()
    return exchange_df


# 不能更改original_result格式（列的分布）
def save_xlsx_result(df):
    # 保存数据
    workbook = Workbook()  # 创建一个新的工作簿
    sheet = workbook.active

    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 15
    for col in range(3, 23):  # 设置第三列到第二十二列的默认列宽为10
        sheet.column_dimensions[get_column_letter(col)].width = 10
    sheet.row_dimensions[1].height = 40  # 设置第一行行高为40

    # 定义单元格填充颜色
    light_gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    gray_fill = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")

    # 将 DataFrame 写入到工作表中
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)

            # 设置第一行自动换行
            if r_idx == 1:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

            # 根据列号设置格式
            if 1 <= c_idx <= 2:  # 第二列为日期格式
                cell.number_format = 'yyyy-mm-dd'
            elif 3 <= c_idx <= 11:  # 第三到第六列为百分比格式
                cell.number_format = '0.0%'
                if 3 <= c_idx <= 6:
                    cell.fill = light_gray_fill  # 浅灰色填充
                elif 7 <= c_idx <= 11:
                    cell.fill = gray_fill
            elif 12 <= c_idx <= 13:
                cell.number_format = '0.00'
            elif 14 <= c_idx <= 23:
                cell.number_format = '0.0%'
                if 14 <= c_idx <= 18:
                    cell.fill = light_gray_fill  # 浅灰色填充
                elif 19 <= c_idx <= 23:
                    cell.fill = gray_fill
    workbook.save(result_file_path + fr"\{fund_code}-brinson归因结果.xlsx")
    print('已成功保存文件！')


if __name__ == '__main__':
    wcodes = pd.read_excel(result_file_path + r'\wcode.xlsx')
    for i in range(len(wcodes)):
        fund_code = wcodes.loc[i, 'fund_code']
        bench_code = wcodes.loc[i, 'bench_code']

        # 对每个fund_code, bench_code进行归因
        # region 基金基准对应数据读取
        # -------------------------数据读取-------------------
        # 获取基金持股权重数据, 主要是利用该数据提取日期区间
        study_fund = gdt.stock_weight(start_dt=start_date, end_dt=end_date, wcode=fund_code, stocks_ind=stocks_ind,
                                      file_path=fund_weight_data_path, fund=True)

        # 获取基准持股权重数据, 权重之和接近1
        bench_fund = gdt.stock_weight(start_dt=start_date, end_dt=end_date, wcode=bench_code, stocks_ind=stocks_ind,
                                      file_path=bench_weight_data_path, bench=True)

        # 基金每日收益率
        fund_ret = gdt.get_ret_daily(start_dt=start_date, end_dt=end_date, wcode=fund_code, file_path=daily_ret_data_path,
                                     fund=True) / 100

        # 指数半年期收益率数据
        index_ret_semi = gdt.get_ret_semiyear(start_dt=start_date, end_dt=end_date, wcode=bench_code,
                                              file_path=semiyear_ret_data_path, bench=True) / 100

        # 基金半年期收益率数据
        fund_ret_semi = gdt.get_ret_semiyear(start_dt=start_date, end_dt=end_date, wcode=fund_code,
                                             file_path=semiyear_ret_data_path, fund=True) / 100
        # --------------------读取完毕-------------------------------------
        # endregion

        # region 归因
        # 单期归因结果, 不调整
        original_result = mulitP_ori(study_fund, bench_fund, stocks_ret, fund_ret_semi, index_ret_semi)

        # 基金换手率
        exchange_rt = exchange_rate(study_fund, stocks_ret).reset_index()
        exchange_rt['start_dt'] = exchange_rt['start_dt'].astype('object')
        original_result = pd.merge(original_result, exchange_rt, on=['start_dt'], how='outer')

        # 基金与各种指数之间的回归后的R2
        r2 = regression_r2(study_fund, fund_ret, bench_ret)
        original_result = pd.merge(original_result, r2, on=['start_dt', 'end_dt'], how='outer')

        # 基金与指数规划求解的结果
        solver_beta1, solver_beta2 = multi_beta(study_fund, fund_ret, bench_ret)
        original_result = pd.merge(pd.merge(original_result, solver_beta1, on=['start_dt', 'end_dt'], suffixes=('_r2', '_beta'),
                                            how='outer'), solver_beta2, on=['start_dt', 'end_dt'], suffixes=('_r2', '_beta'), how='outer')
        # endregion

        print(original_result)
        save_xlsx_result(original_result)




