import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import warnings
from openpyxl import load_workbook
warnings.filterwarnings("ignore")
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False  # 如果负号不显示，改为 True

# 需要改的参数
file_path = r'F:\TianMChun\intership\shouchuang\work\行业轮动\基于拥挤度匹配行业轮动策略\data'
last_month_day = 22  # 最后一个月交易日数量
excel_file_name = r'\基于拥挤度匹配行业轮动策略月报模版.xlsx'



# 模型固定参数
mkt_weight_label = 0  # 根号市值对主成分分析之前的收益矩阵加权，1是log市值加权，2是不加权
pca_explained_ratio = 0.9  # 主成分累计解释度
period_window = 1   # period=window/period_window,半衰期与采样窗口的关系
window = 60  # 采样窗口长度，3个月
significant_return = 0.095
exchange_fee = 0.0025  # 交易成本
history_sample_st = '2018-01-01'  # 历史样本开始日期，计算当前集中度序列与历史集中度序列的相似度
similarity_st = '2020-11-01'  # 相似度计算开始日期
start_nv_date = '2021-01-01'  # 样本开始日期，2021年

# <editor-fold desc="数据读取">
# ------------------数据读取(index为Date, columns为中信行业)-------------------------
data_open = pd.read_excel(file_path + r'\ZXOpen.xlsx').dropna()
data_open['Date'] = data_open['Date'].astype(int).astype(str)
data_open['Date'] = pd.to_datetime(data_open['Date'], format="%Y%m%d")
data_open.set_index(data_open.columns[0], inplace=True)


data_close = pd.read_excel(file_path + r'\ZXPrice.xlsx').dropna()
data_close['Date'] = data_close['Date'].astype(int).astype(str)
data_close['Date'] = pd.to_datetime(data_close['Date'], format="%Y%m%d")
data_close.set_index(data_close.columns[0], inplace=True)


mkt_size = pd.read_excel(file_path + r'\ZXSizeT.xlsx').dropna()
mkt_size['Date'] = mkt_size['Date'].astype(int).astype(str)
mkt_size['Date'] = pd.to_datetime(mkt_size['Date'], format="%Y%m%d")
mkt_size.set_index('Date', inplace=True)
# </editor-fold>
# --------------------------数据读取完毕--------------------------------------


# 生成window个递增的权重
def get_half_life_weight(half_life, window):
    """
    生成指数衰减的递减权重列表，半衰期为half_life，权重个数为window
    """
    lamda = np.log(2) / half_life  # 衰减率
    return (np.exp(-lamda * np.arange(window)))[::-1]


def centrality_single(df, explained_ratio):
    """
    :param df: 已加权日度收益数据，array
    :param explained_ratio: 主成分分析解释度
    :return: 每个行业的集中度
    """
    df_standard = (df - df.mean(axis=0)) / df.std(axis=0)  # N*30  -mean/std
    df_standard = np.nan_to_num(df_standard, nan=0)
    scaled_cov = np.cov(df_standard, rowvar=False)

    # 求解释度>=explained_ratio的特征值，特征向量
    eigenvalues, eigenvectors = np.linalg.eig(scaled_cov)  # 特征向量是一列一个  默认特征值降序
    explained_variance_ratio = eigenvalues / eigenvalues.sum()
    n = np.argmax(np.cumsum(explained_variance_ratio) >= explained_ratio) + 1
    eigenvalues = eigenvalues[:n]  # shape=(n,)
    eigenvectors = eigenvectors[:, :n]  # shape=(30, n)
    
    # 吸收比率
    absorption_ratio = eigenvalues / pd.DataFrame(df).var().sum()

    # 集中度
    eigenvectors_abs = np.abs(eigenvectors)
    eigenvectors_abs_wgt = eigenvectors_abs / eigenvectors_abs.sum(axis=0)  
    centrality_ = np.sum(absorption_ratio * eigenvectors_abs_wgt, axis=1) / absorption_ratio.sum()

    return centrality_


# 动态求集中度
def centrality_dynamic(df_close, half_life, window, df_mkt, explained_ratio):
    # 不同市值加权方式的标记
    global mkt_weight_label

    # 根据收盘价计算每日收益数据
    df_pct = df_close.pct_change()
    df_pct = df_pct.replace([np.inf, -np.inf], np.nan).dropna()
    df_pct = df_pct[df_pct.index >= '2017-01-01']  # 只计算2017年开始的集中度

    # 半衰期权重
    half_life_weight = get_half_life_weight(half_life, window)

    # 初始化一个空的 numpy 数组来存储结果
    centrality_list = []

    # 动态计算行业集中度
    # 计算t=T的的集中度，需要t=T-window开始到t=T的数据
    for i in range(window, len(df_pct)):
        # 原始收益矩阵
        sub_array = df_pct.iloc[i-window: i, :]

        # 市值加权
        weight_mkt = df_mkt[df_mkt.index.isin(df_pct.index[i - window: i])]  # 对应日期的市值数据

        # 根号市值加权
        if mkt_weight_label == 0:
            mkt_weight = (weight_mkt.astype('float') ** 0.5)
            # mkt_weight = (weight_mkt.astype('float') ** 0.5) / (weight_mkt.astype('float') ** 0.5).sum()  # 应该是这样的，但是报告里是上一行
            sub_array = sub_array * mkt_weight

        # 对数市值加权
        if mkt_weight_label == 1:
            weight_mkt[weight_mkt == 0] = 1e-10  # 防止log之后出现inf
            mkt_weight = np.log(weight_mkt.astype('float'))
            # mkt_weight = np.log(weight_mkt.astype('float')) / np.log(weight_mkt.astype('float')).sum()
            sub_array = sub_array * mkt_weight

        # 不用市值加权
        if mkt_weight_label == 2:
            pass

        sub_array = sub_array.multiply(half_life_weight, axis=0)  # (window,30)
        sub_array = sub_array.fillna(0)
        centrality_ = centrality_single(sub_array.values, explained_ratio)
        centrality_list.append(centrality_)

    # 将结果转换回 DataFrame
    result = pd.DataFrame(centrality_list, index=df_pct.index[window:], columns=df_pct.columns)
    return result


# 相似度
def up_down_pct(cen_df, df_close, history_start_dt, similarity_start_dt, ret):
    # 最后一个月交易日数量
    global last_month_day

    # 只需要>start_dt的数据
    cen_df = cen_df[cen_df.index > history_start_dt]
    df_close = df_close[df_close.index > history_start_dt]

    # 提取月度日期 月末
    cen_df['year_month'] = cen_df.index.to_period('M')
    mon_index = cen_df.drop_duplicates(subset=['year_month'], keep='last').index
    cen_df.drop('year_month', axis=1, inplace=True)

    # 样本月度日期
    dt_mon_end = mon_index[mon_index >= similarity_start_dt]

    # 日度日期
    day_index = cen_df.index.to_numpy()

    # 相似度 DataFrame
    df_result = pd.DataFrame(
        index=pd.MultiIndex.from_product([dt_mon_end, [ret]], names=['dt_mon_begin', 'divide_ret']),
        columns=cen_df.columns)

    ## 统计显著上涨下跌样本数量
    df_up = pd.DataFrame(index=dt_mon_end, columns=cen_df.columns)
    df_down = pd.DataFrame(index=dt_mon_end, columns=cen_df.columns)
    ##

    # 遍历每个月计算相似度
    for i, dt in enumerate(dt_mon_end):
        # 历史数据的收盘价
        df_close_past = df_close[df_close.index <= dt]

        # 当前月和下月交易日数量
        n_now_index = day_index[(day_index <= dt) & (day_index > mon_index[mon_index.get_loc(dt)-1])]
        n_now = len(n_now_index)  # 当月交易日数量
        if i < len(dt_mon_end) - 1:
            n_next = len(day_index[(day_index <= dt_mon_end[i + 1]) & (day_index > dt)])
        else:
            n_next = last_month_day

        # 计算历史数据收益率
        rolling_window = n_now + n_next
        df_past_ret = df_close_past.rolling(window=rolling_window).apply(
            lambda x: x[rolling_window-1] / x[n_now-1] - 1).shift(-(rolling_window - 1))
        df_past_ret = df_past_ret.apply(lambda row: row.fillna(row.median()), axis=1)

        # 显著上涨或下跌标签，上涨标签为1，下跌标签为-1
        df_past_ret_label = pd.DataFrame(
            np.where(df_past_ret > ret, 1, np.where(df_past_ret < -ret, -1, np.nan)),
            index=df_past_ret.index,
            columns=df_past_ret.columns
        )

        ## 统计显著上涨下跌样本数量
        df_up.loc[dt] = df_past_ret_label[df_past_ret_label > 0].count()
        df_down.loc[dt] = df_past_ret_label[df_past_ret_label < 0].count()

        # 若df_past_ret_label不为nan，则在df_close_past对应位置往后提取n_next行数据，与df_cen_now对应列计算残差平方和，得到新的df_resid
        df_distance = pd.DataFrame(np.nan, index=df_past_ret.index, columns=df_past_ret.columns)
        now_cen = cen_df.loc[n_now_index, :]
        past_day_index = df_past_ret.dropna().index
        for j, idx in enumerate(past_day_index):
            # 找到当前行中非 NaN 的列
            valid_cols = df_past_ret_label.columns[~df_past_ret_label.loc[idx].isna()]

            # 提取历史样本集中度序列和当前日期集中度序列
            history_cen = cen_df.loc[idx:day_index[j+n_now-1], valid_cols].to_numpy()
            now_cem_col = now_cen[valid_cols].to_numpy()

            # 计算残差平方和
            resid = np.sqrt(np.sum((history_cen - now_cem_col) ** 2, axis=0))
            df_distance.loc[idx, valid_cols] = resid

        # 使用历史收益对距离加权
        df_past_ret_signi = np.sqrt((df_past_ret * df_past_ret_label).abs())
        weight_ = df_past_ret_signi / df_past_ret_signi.sum()
        df_distance_weighted = df_distance * weight_

        weight_ = np.sqrt(df_past_ret.abs() / (df_past_ret * df_past_ret_label).abs().sum())
        df_distance_weighted = df_distance * weight_

        # 下跌的欧氏距离除以 上涨+下跌的欧氏距离
        down_mean = df_distance_weighted[df_past_ret_label < 0].sum()
        up_mean = df_distance_weighted[df_past_ret_label > 0].sum()
        df_result.loc[(dt, ret)] = down_mean / (down_mean + up_mean)
    return df_result


def format_DaytoMon(df):
    """
    输入一个日频df（index为日期， columns为30个行业， values为centrality），
    返回月频双重索引df(第一层索引为date，第二重索引为industry,value为centrality)
    """
    df['year_month'] = df.index.to_period('M')
    df = df.drop_duplicates(subset=['year_month'], keep='last')
    df.drop('year_month', axis=1, inplace=True)

    # 堆叠，形成双重索引，第一层是date, 第二层是industry
    df_stack = pd.DataFrame(df.stack())
    df_stack.index.names = ['date', 'industry']
    df_stack.columns = ['centrality']
    return df_stack


def group_factor(factor_df, q=5):
    """
    对双重索引因子df分组, 组1因子值最大
    :param factor_df: 双重索引，value为因子值，只能有一列因子值
    :param q: 一共分为q组
    :return: 双重索引df，value为因子值和分组结果
    """
    f_grouped = factor_df.groupby('date').rank(ascending=False)  # 降序排序，组1最大
    f_grouped.columns = ['rank']
    f_grouped = pd.merge(f_grouped.reset_index(),
                         f_grouped.groupby('date')['rank'].count().to_frame('count'), how='left', on='date')
    f_grouped['group'] = np.ceil(f_grouped['rank'] / (f_grouped['count'] * (1 / q)))
    f_grouped = f_grouped.set_index(['date', 'industry'])
    f_grouped = pd.concat([factor_df, f_grouped], axis=1)
    return f_grouped


def sample_return(factor_group, df_close, df_open, start_date):
    """
    计算已分组的因子df, 考虑成本的每组收益
    :param start_date: 因子收益计算开始日期
    :param df_open: 开盘价df, index为date, columns为industry, value为开盘价
    :param factor_group:  已分组因子df
    :param df_close: 收盘价df, index为date, columns为industry, value为收盘价
    :return: 基准收益，每组收益，每组超额收益，组成的df
    """
    # 获取所有的月频，日频交易日期
    factor_dates = factor_group.index.get_level_values(0).unique()
    monthly_date = factor_dates.to_list()  # 月度交易日期
    daily_dates = df_close.index.to_list()

    # 行业收益，日频
    ret_df = pd.DataFrame(index=df_close.index[df_close.index >= start_date])

    # 原始收益
    close_pct = df_close.pct_change().replace(np.nan, 0)

    # 开始计算收益日期后的月度日期
    filter_date = factor_dates[factor_dates >= start_date].to_list()

    # 交易成本，换手率，所选行业
    trade_cost = pd.DataFrame(index=filter_date)
    exchange_rate = pd.DataFrame(index=filter_date)
    book_id_df = pd.DataFrame(index=filter_date, columns=['book_id'])

    for g in range(1, int(factor_group['group'].max()) + 1):
        for idx in range(len(filter_date)):
            end_dt = filter_date[idx]  # 本月末日期
            last_end_dt = monthly_date[monthly_date.index(end_dt) - 1]  # 上月末日期

            start_dt_1 = daily_dates[daily_dates.index(last_end_dt) + 1]  # 获取每月第一个交易日日期
            start_dt_2 = daily_dates[daily_dates.index(last_end_dt) + 2]  # 获取每月第二个交易日日期

            book_id = list(factor_group[factor_group['group'] == g].loc[last_end_dt,].index)

            # 本月初第二天到本月末的原始收益 = 当天收盘 / 上一天收盘 - 1
            group_pct = close_pct.loc[start_dt_2: end_dt, book_id]

            # 第一个月的收益
            if idx == 0:
                # 本月初第一天的原始收益 = 本月初第一天收盘 / 上月末最后一天收盘 -1 ,没有跳空收益
                group_pct.loc[start_dt_1] = df_close.loc[start_dt_1, book_id].astype(float) / df_close.loc[last_end_dt, book_id].astype(float) - 1
                group_pct = group_pct.sort_index()

                # 换手率
                exchange_rate_ = 1

            else:
                last_last_end_dt = monthly_date[monthly_date.index(end_dt) - 2]  # 上上月末日期
                # print(end_dt, last_end_dt, last_last_end_dt)

                last_start_1 = df_close.index[df_close.index.get_loc(last_last_end_dt) + 1]  # 获取上月第一个交易日日期
                # print(start_dt_1, start_dt_2, last_start_1)

                book_id_last = list(factor_group[factor_group['group'] == g].loc[last_last_end_dt,].index)

                # 换手率
                ori_weight = close_pct.loc[last_start_1:last_end_dt, book_id_last].add(1).cumprod()
                ori_weight = ori_weight.iloc[-1].transform(lambda x: x / x.sum())
                new_weight = pd.Series([1 / 6] * 6, index=book_id)
                weights = pd.concat([ori_weight, new_weight], axis=1).fillna(0)
                exchange_rate_ = (weights.iloc[:, 0] - weights.iloc[:, 1]).abs().sum()

                # 本月初第一天的原始收益1 = 本月初第一天收盘 / 本月初第一天开盘 -1
                group_pct.loc[start_dt_1] = df_close.loc[start_dt_1, book_id].astype(float) / df_open.loc[start_dt_1, book_id].astype(float) - 1
                group_pct = group_pct.sort_index()

                # 本月初第一天的原始收益2（跳空收益） = (本月初第一天开盘 - 上月末收盘) / 上月初第一天开盘
                open_ret = (df_open.loc[start_dt_1, book_id_last].astype(float) / df_open.loc[last_start_1, book_id_last].astype(float)).mean()
                close_ret = (df_close.loc[last_end_dt, book_id_last].astype(float) / df_open.loc[last_start_1, book_id_last].astype(float)).mean()
                group_pct2 = open_ret / close_ret - 1

                # 第一天的收益
                group_pct.loc[start_dt_1] = (group_pct.loc[start_dt_1] + 1) * (group_pct2 + 1) - 1

            # 换手率
            exchange_rate.loc[end_dt, f'组{g}'] = exchange_rate_

            # 交易成本
            cost = exchange_rate_ / 2 * exchange_fee
            trade_cost.loc[end_dt, f'组{g}'] = cost

            # 所选行业
            book_id_df.loc[end_dt, 'book_id'] = book_id

            # 分月度计算基准的累计收益
            bench_cum_ret = (close_pct.loc[start_dt_1:end_dt,].add(1).cumprod()-1).mean(axis=1)
            ret_df.loc[start_dt_2:end_dt, '基准'] = bench_cum_ret.add(1).pct_change().dropna()
            ret_df.loc[start_dt_1, '基准'] = bench_cum_ret.loc[start_dt_1]

            # 每组收益
            group_pct.loc[start_dt_1] = group_pct.loc[start_dt_1] - cost  # 交易成本
            group_cum_ret = (group_pct.add(1).cumprod() - 1).mean(axis=1)
            ret_df.loc[start_dt_2:end_dt, f'组{g}'] = group_cum_ret.add(1).pct_change().dropna()
            ret_df.loc[start_dt_1, f'组{g}'] = group_cum_ret.loc[start_dt_1]

            # 超额收益
            excess_cum_ret = group_cum_ret - bench_cum_ret
            ret_df.loc[start_dt_2: end_dt, f'组{g}超额'] = excess_cum_ret.add(1).pct_change().dropna()
            ret_df.loc[start_dt_1, f'组{g}超额'] = excess_cum_ret.loc[start_dt_1]
    return ret_df


def next_month_rank(factor_group, df_close):
    """
    :param factor_group: group_factor函数的返回结果，分好组的因子值，30个行业 五组 每组6个行业
    :param df_close: idx为日期的close
    :return: rank_next：当月选出的行业下月的排名，组一（排名靠前）的所有行业下期close收益率排名
    """
    # df_close = data_close
    # factor_group = centrality_group
    df_close_m = df_close[df_close.index.isin(factor_group.index.get_level_values(0))]
    df_close_m_pct = df_close_m.pct_change().dropna()
    rank_pct = df_close_m_pct.rank(axis=1, ascending=False)  # 对每月收益降序排序，收益最高，排名第一

    # 每月选出的行业
    rank_now = pd.pivot(factor_group.reset_index(), index='date', columns='industry', values='rank')
    rank_now = rank_now[rank_now <= 6]

    # 当月行业下期排名
    rank_next = (rank_pct.shift(-1) * rank_now) / rank_now

    return rank_next, rank_now


if __name__ == '__main__':

    # 月度收益，计算IC需要
    data_pct_m = data_close.copy()
    data_pct_m['year_month'] = data_pct_m.index.to_period('M')
    data_pct_m = data_pct_m.drop_duplicates(subset=['year_month'], keep='last')
    data_pct_m.drop('year_month', axis=1, inplace=True)
    data_pct_m = data_pct_m.pct_change()

    # 月频交易日，月末
    month_dt = data_pct_m.index

    centrality_df = centrality_dynamic(data_close, half_life=window/period_window, window=window, df_mkt=mkt_size, explained_ratio=pca_explained_ratio)

    # 与历史样本未来一个月显著上涨下跌的相似度
    similar_df = up_down_pct(centrality_df, data_close, history_sample_st, similarity_st, significant_return)
    similar_df = similar_df.applymap(lambda x: x.real)

    # 把月初日期换成月末日期，便于后面计算nv
    similar_df_ = similar_df.xs(significant_return, level=1)
    similar_df_.index = month_dt[month_dt >= similarity_st]

    # 堆叠 双重索引
    centrality_df_stack = pd.DataFrame(similar_df_.stack())
    centrality_df_stack.index.names = ['date', 'industry']
    centrality_df_stack.columns = ['centrality']

    # IC
    df = pd.concat([centrality_df_stack, data_pct_m.shift(-1).loc[similarity_st:,].stack()], axis=1)
    ic = df.groupby(level=0).apply(lambda x: x.corr('spearman').iloc[0, 1])

    # 集中度分组
    centrality_group = group_factor(centrality_df_stack, q=5)

    # 计算每组收益
    df_ret = sample_return(centrality_group, data_close, data_open, start_nv_date)

    # 计算当期行业下期排名
    rank_next_month, rank_this_month = next_month_rank(centrality_group, data_close)


    # <editor-fold desc="更新excel">
    excel_file_path = file_path + excel_file_name
    wb = load_workbook(excel_file_path)

    # 修改sheet '最新组合'
    ws_new_group = wb['最新组合']

    # 所需数据， 相似度日期往后推一个月，意味着当前月预测未来一个月的相似度
    similar_df_shift = similar_df_.copy()
    similar_df_shift.index = [idx for idx in similar_df_.index - pd.DateOffset(months=-1)]
    last_two_rows = similar_df_shift.iloc[-2:].T.values  # 获取 similar_df_shift 最后两行的数据,即最新两个月的数据

    # 写入 F2:G31，即最新两个月的数据
    start_row, start_col = 2, 6  # F2 起始位置
    for i, row_values in enumerate(last_two_rows, start=start_row):
        for j, value in enumerate(row_values, start=start_col):
            ws_new_group.cell(row=i, column=j, value=value)


    # 修改sheet '上月跟踪'
    ws_past_trend = wb['上月跟踪']

    # 所需数据
    ranked_values = data_pct_m.rank(axis=1, ascending=False).iloc[-1, :].T.values  # 计算 rank 并提取最后一行的排名值

    # 写入 C2:C30，即实际收益排名
    start_row, col = 2, 3  # C2 起始位置
    for i, value in enumerate(ranked_values, start=start_row):
        ws_past_trend.cell(row=i, column=col, value=value)


    # 修改sheet '今年收益'
    ws_this_year_ret = wb['今年收益']

    # 所需数据
    df_ret_this_year = df_ret[df_ret.index.year == df_ret.index.year.max()]
    df_ret_this_year['替换日期'] = df_ret_this_year.index.strftime('%Y%m%d')
    df_need = df_ret_this_year[['替换日期', '组1', '基准', '组1超额']].values

    # 删除B:E 列第2行开始后的旧数据
    for row in ws_this_year_ret.iter_rows(min_row=2, min_col=2, max_col=5):
        for cell in row:
            cell.value = None

    # 写入 B:E 列第2行开始后的数据
    start_row, start_col = 2, 2  # B 列对应列号 2，起始行为 2
    for i, row_data in enumerate(df_need, start=start_row):
        for j, value in enumerate(row_data, start=start_col):  # 写入 B 到 E 列
            ws_this_year_ret.cell(row=i, column=j, value=value)

    # 所需IC数据
    ic_need = ic.copy()
    ic_need.index = [idx for idx in ic_need.index - pd.DateOffset(months=-1)]
    ic_need = ic_need[ic_need.index.year == ic_need.index.year.max()]
    ic_need.index = ic_need.index.strftime("%Y年%m月")
    ic_need.name = 'IC'
    ic_need = ic_need.reset_index().dropna().values

    # 删除 S2:T13 区域的旧数据
    for row in ws_this_year_ret.iter_rows(min_row=2, max_row=13, min_col=19, max_col=20):
        for cell in row:
            cell.value = None

    # 写入 S2:T13 区域的数据
    start_row, start_col = 2, 19  # S2 对应行号 2，列号 19 (S列)
    for i, row_data in enumerate(ic_need, start=start_row):  # 限制到前 12 行 (S2:T13)
        for j, value in enumerate(row_data, start=start_col):  # 只填充 S 列和 T 列 (2 列数据)
            ws_this_year_ret.cell(row=i, column=j, value=value)


    # 更新sheet '今年排名'
    ws_this_year_rank = wb['今年排名']

    # 所需排名数据
    df_rank_this_year = rank_next_month.copy()
    df_rank_this_year.index = df_rank_this_year.index[1:].append(pd.Index([None]))  # 直接将index的顺序向上移动一行
    df_rank_this_year = df_rank_this_year[df_rank_this_year.index.year == df_rank_this_year.index.year.max()].T.values  # 只选择今年的数据

    # 删除C2:M31 区域的排名数据
    for row in ws_this_year_rank.iter_rows(min_row=2, max_row=31, min_col=3, max_col=13):
        for cell in row:
            cell.value = None

    # 写入 C2:M31 区域的排名数据
    start_row, start_col = 2, 3  # C3 对应行号 3，列号 3 (C 列)
    for i, row_data in enumerate(df_rank_this_year, start=start_row):
        for j, value in enumerate(row_data, start=start_col):
            ws_this_year_rank.cell(row=i, column=j, value=value)


    # 更新sheet'历史收益'
    ws_history_ret = wb['历史收益']

    # 所需数据
    df_ret_history = df_ret.copy()
    df_ret_history['替换日期'] = df_ret_history.index.strftime('%Y%m%d')
    df_need_history = df_ret_history[['替换日期', '组1', '基准', '组1超额']].values

    # 删除B:E 列第2行开始后的旧数据
    for row in ws_history_ret.iter_rows(min_row=2, min_col=2, max_col=5):
        for cell in row:
            cell.value = None

    # 写入 B:E 列第2行开始后的数据
    start_row, start_col = 2, 2  # B 列对应列号 2，起始行为 2
    for i, row_data in enumerate(df_need_history, start=start_row):
        for j, value in enumerate(row_data, start=start_col):  # 写入 B 到 E 列
            ws_history_ret.cell(row=i, column=j, value=value)

    # 查看工作表数据
    # data = ws_this_year_ret.values
    # columns = next(data)  # 获取第一行作为列名（如果有）
    # df = pd.DataFrame(data, columns=columns)
    # print(df)
    # 保存修改后的工作簿
    wb.save(excel_file_path)
    # </editor-fold>



